import openpyxl
def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)


def getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readrow(file,sheetname,rownum,columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=columnum).value

def writecolumn(file,sheetname,rownum,columnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnum).value=data
    workbook.save(file)